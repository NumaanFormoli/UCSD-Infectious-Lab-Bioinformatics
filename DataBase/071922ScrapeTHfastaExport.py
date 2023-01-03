import openpyxl
import os
import sys
from datetime import date

#todo rechcek mutants for Ns - 44 Ns
#todo lookup blanks to see
#todo flesh out duplicate error checker

#dictionary to hold sequences, a counter in case of duplicates, parent directory so python knows where to look, and fasta file
mutdict = {}
dupecounter = 0
pdir = os.getcwd() + "/"
fafile = "Database/EPI_ISL_1167892.fasta"
#opens the fasta file as literally a text file
with open(pdir + fafile) as file:
    #lastmut used to check when onto next mutation - updates when it hits a ">"
    lastmut = ""
    #for loop goes one new line at a time, sline is stripped just in case of wonky format. (strip and dup checking techincally optional if file is clean)
    for line in file:
        sline = line.strip()
        if sline[0] == ">":
            #quick error catching in case names are the same - adds copy if so, otherwise creates a dict key
            #either way lastmut is updated and creates a new key in mutdict which is "" blank for now
            if sline in mutdict:
                mutdict[sline + "_" + str(dupecounter)] = ""
                lastmut = sline + "_" + str(dupecounter)
                dupecounter +=1
            else:
                mutdict[sline] = ""
                lastmut = sline
        #should trigger on every other line - those lines should be RMSAX-MAN..., and then assigned to mutdict of last mutation. += appends so the sequence continues until it hits a new ">"
        #should throw error if file doesn't start with ">name"
        else:
            mutdict[lastmut] += sline
#here actually creating the excel file to populate with the sequences. wb is the excel file itself, and sh1 is the first sheet.
wb = openpyxl.Workbook()
sh1 = wb.active
#row counter to keep track as for loop goes down. Starts at 2 to give room for titles in first row
rowcount = 2
sh1.cell(1, 1).value = "Seq"
sh1.cell(1, 2).value = "Name"
#seq goes first in column 1, just because how jupyter file is built. Syntax is cell(row, column). Assigning value to the seq in mut dict and then the name/key (x)
for x in mutdict.keys():
    sh1.cell(rowcount, 1).value = mutdict[x]
    sh1.cell(rowcount, 2).value = x
    rowcount +=1
#Saving/exporting the xls file. can change to xlsx depnding on how pandas's error is feeling in the jupyter file. date added for GDP
wb.save("Practice_new_nprotein_seq.xlsx")
sys.exit("Finished with " + str(dupecounter) + " duplicates")
